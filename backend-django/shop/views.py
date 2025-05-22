from django.shortcuts import render, get_object_or_404
from rest_framework.views import APIView
from .models import Course
from rest_framework.response import Response
from . serializers import *
import openpyxl
from django.http import HttpResponse
from io import BytesIO
from django.contrib.auth.models import User

# Create your views here.


class ReactView_Course(APIView):
    serializer_course = CourseSerializer

    def get(self, request):
        detail = [{"id": detail.id, "title": detail.title, "price": detail.price,
                   "category": str(detail.category),
                   "students_qty": detail.students_qty,
                   "reviews_qty": detail.reviews_qty, }
                  for detail in Course.objects.all()]
        return Response(detail)

    def post(self, request):

        serializer = CourseSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save
            return Response(serializer.data)


def export_courses_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Курсы"

    headers = ['ID', 'Название', 'Категория', 'Цена']
    sheet.append(headers)

    for course in Course.objects.all():
        sheet.append([
            course.id,
            course.title,
            course.category.title if course.category else '',
            course.price
        ])

    sheet_categories = workbook.create_sheet(title="Категории")
    headers_categories = ['ID', 'Название', 'Создана']
    sheet_categories.append(headers_categories)

    for category in Category.objects.all():
        sheet_categories.append([
            category.id,
            category.title,
            category.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    sheet_users = workbook.create_sheet(title="Пользователи")
    headers_users = ['ID', 'Имя пользователя', 'Email']
    sheet_users.append(headers_users)

    for user in User.objects.all():
        sheet_users.append([
            user.id,
            user.username,
            user.email
        ])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="courses.xlsx"'
    return response
